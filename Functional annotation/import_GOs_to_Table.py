import openpyxl
import requests
import argparse
from requests.exceptions import SSLError
import psutil
import os
import time
import threading
import requests
from requests.exceptions import SSLError, Timeout

def main():

    welcome_header = """
    ******************************************************
    *                                                    *
    *                Welcome to import_GO.py             *
    *                                                    *
    ******************************************************
    """

    script_name = "import_GO.py"
    creator = "Gabriel Dall'Alba"
    date = "2024-02-20"
    description = "This script imports GO terms from an Excel file and retrieves their names using an API."
    usage = "python import_GO.py <file>"

    print(welcome_header)

    print(f"Script Name: {script_name}")
    print(f"Creator: {creator}")
    print(f"Date: {date}")
    print(f"Description: {description}")
    print(f"Usage: {usage}")

    parser = argparse.ArgumentParser(description="Import GO terms from an Excel file")
    parser.add_argument("file", help="Path to the Excel file")
    #parser.add_argument("output", help="Path to the output file")
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.file)
    sheet = workbook.active

    number_of_GOs_total = 0
    number_of_GOs = 0

    # Dictionary to store the GO terms and their names
    go_terms_dict = {}
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and cell.value.startswith("GO"):
                number_of_GOs_total = number_of_GOs_total + 1

    print(f"Number of GOs: {number_of_GOs_total}")

    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell contains a GO term
            if cell.value and cell.value.startswith("GO:"):
                go_term = cell.value
                number_of_GOs = number_of_GOs + 1
                #if number_of_GOs % 100 == 0:
                    #print(f"Processed {number_of_GOs} GO terms out of {number_of_GOs_total}")

                # Check if the GO term is already in the dictionary
                if go_term in go_terms_dict:
                    go_term_name = go_terms_dict[go_term]
                else:
                    # Request the name of the GO term through the API
                    go_term_name = get_go_term_name(go_term)
                    go_terms_dict[go_term] = go_term_name

                # Count the number of substitutions
                substitution_count = 0

                # Search and substitute the GO term in the entire Excel sheet
                for search_row in sheet.iter_rows():
                    for search_cell in search_row:
                        if search_cell.value == go_term:
                            search_cell.value = f"{go_term_name} ({go_term})"
                            substitution_count += 1

                #print(f"Substituted {substitution_count} instances of {go_term} with {go_term_name}")
                # Print the number of substitutions in each iteration
                            

                # Save the GO term and its name to the input file
                workbook.save(args.file)

#                if go_term_name:  # Ensure there's a name before writing
#                    print(f"Modified cell value: {go_term} -> {go_term_name}")

#                else:
#                    print(f"Error retrieving name for {go_term}")

    # Close the Excel file
    workbook.close()


def get_go_term_name(go_term):
    try:
        # Make a request to the API to get the name of the GO term
        url = f"https://www.ebi.ac.uk/QuickGO/services/ontology/go/terms/{go_term}"
        response = requests.get(url, timeout=10)
        response.raise_for_status()  # Raises a HTTPError if the status is 4xx, 5xx

        if response.status_code == 200:
            data = response.json()
            #print(f"Found {go_term} in the API.")
            return data['results'][0]['name']
        else:
            print(f"Failed to get data from API for GO term: {go_term}")
            return None

    except SSLError:
        #print(f"SSL Error for GO term: {go_term}. Retrying with SSL verification disabled.")
        response = requests.get(url, verify=False, timeout=5)
        if response.status_code == 200:
            data = response.json()
            #print(f"Found {go_term} in the API.")
            return data['results'][0]['name']
        else:
            print(f"Failed to get data from API for GO term with SSL disabled: {go_term}")
            return None

    except Timeout:
        print(f"Timeout error for GO term: {go_term}.")
        return None

    except requests.exceptions.RequestException as e:
        print(f"Error occurred for GO term: {go_term}. {str(e)}")
        return None
    

    
if __name__ == "__main__":
    main()
